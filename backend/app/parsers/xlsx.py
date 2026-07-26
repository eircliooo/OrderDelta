"""XLSX 解析器。SPEC §5.2。

**加载协议（踩坑清单 #1-#3，照做，不要"优化"）：**
  pass A  data_only=False  取公式串、merged_cells.ranges、地址、data_type
  pass B  data_only=True   只取缓存值
  两次都禁止 read_only=True —— read_only 下 merged_cells 不可用，
  合并大标题定位会失效。

  FORMULA_WITHOUT_CACHE = pass A 是公式 **且** pass B 为 None。
  别把「值为空」误报成「无缓存」。

  行数/表数限制**边遍历边判**，不要读 <dimension> 预检 —— openpyxl 官方警告
  该值常被写错，恶意文件声明 A1:A1 即可绕过。

解析器只输出**原始网格**。表头定位是提取层的事（extraction/header.py），
所以 ParsedTable.header_rows 在这里恒为空元组。
"""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.enums import ParseReasonCode, ParseStatus
from app.parsers.base import (
    DEFAULT_LIMITS,
    CellRef,
    CoordinateSpace,
    DocumentInput,
    ParseCapability,
    ParsedCell,
    ParsedDocument,
    ParsedTable,
    ParseLimits,
    TextBlock,
)

PARSER_NAME = "openpyxl-xlsx"
PARSER_VERSION = "1.0.0"

#: 只接受 .xlsx。.xlsm 含宏一律拒绝（不执行宏 = 连打开都不给）。
ACCEPTED_SUFFIXES = frozenset({".xlsx"})

_ZIP_MAGIC = b"PK\x03\x04"
#: OLE2 复合文档头：老式 .xls，或**加密的** OOXML 文件。
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

#: ZIP 炸弹防护（SPEC §15.2）。openpyxl 基于 zipfile+lxml，
#: 仅限制上传文件大小挡不住高压缩比的解压炸弹。
MAX_UNCOMPRESSED_BYTES = 300 * 1024 * 1024
MAX_COMPRESSION_RATIO = 200

#: 每个 sheet 输出的 TextBlock 上限（文档级字段提取与 MVP-1 LLM 引用用）。
MAX_BLOCKS_PER_SHEET = 2000


class XlsxParser:
    """实现 `app.parsers.base.DocumentParser` 协议。"""

    name = PARSER_NAME
    version = PARSER_VERSION

    # ------------------------------------------------------------------ can_parse

    def can_parse(self, src: DocumentInput) -> ParseCapability:
        suffix = Path(src.original_filename).suffix.lower()
        if suffix not in ACCEPTED_SUFFIXES:
            if suffix == ".pdf":
                return ParseCapability.reject(
                    ParseReasonCode.UNSUPPORTED_EXT,
                    "MVP-0 暂不支持 PDF，将在下一版本提供",
                )
            if suffix in {".xls", ".xlsm", ".xlsb"}:
                return ParseCapability.reject(
                    ParseReasonCode.UNSUPPORTED_EXT,
                    f"仅支持 .xlsx，不支持 {suffix}（.xlsm 含宏，一律拒绝）",
                )
            return ParseCapability.reject(
                ParseReasonCode.UNSUPPORTED_EXT, f"不支持的文件类型 {suffix or '（无扩展名）'}"
            )

        header = _read_magic(src.path)
        if header.startswith(_OLE2_MAGIC):
            # 扩展名是 .xlsx 但内容是 OLE2 -> 多半是加密的 OOXML
            return ParseCapability.reject(
                ParseReasonCode.ENCRYPTED, "文件已加密或不是真正的 .xlsx，无法读取"
            )
        if not header.startswith(_ZIP_MAGIC):
            return ParseCapability.reject(
                ParseReasonCode.CORRUPT, "文件头不是有效的 xlsx（zip）结构"
            )

        bomb = _check_zip_bomb(src.path)
        if bomb is not None:
            return bomb

        return ParseCapability.ok()

    # ---------------------------------------------------------------------- parse

    def parse(self, src: DocumentInput, limits: ParseLimits = DEFAULT_LIMITS) -> ParsedDocument:
        capability = self.can_parse(src)
        if not capability.accepted:
            return self._failed(ParseStatus.REJECTED, capability.reason_code, capability.detail)

        try:
            wb_formula = load_workbook(src.path, data_only=False, read_only=False)
            wb_values = load_workbook(src.path, data_only=True, read_only=False)
        except zipfile.BadZipFile:
            return self._failed(ParseStatus.FAILED, ParseReasonCode.CORRUPT, "文件损坏")
        except Exception as exc:  # openpyxl 抛型很杂，一律显式失败而不是崩掉请求
            return self._failed(
                ParseStatus.FAILED, ParseReasonCode.CORRUPT, f"无法打开工作簿：{type(exc).__name__}"
            )

        try:
            return self._parse_workbook(wb_formula, wb_values, limits)
        finally:
            wb_formula.close()
            wb_values.close()

    # ------------------------------------------------------------------ internals

    def _failed(
        self,
        status: ParseStatus,
        reason: ParseReasonCode | None,
        detail: str | None,
    ) -> ParsedDocument:
        return ParsedDocument(
            parser_name=self.name,
            parser_version=self.version,
            status=status,
            reason_code=reason,
            detail=detail,
        )

    def _parse_workbook(
        self, wb_formula: Any, wb_values: Any, limits: ParseLimits
    ) -> ParsedDocument:
        sheet_names = list(wb_formula.sheetnames)
        warnings: list[str] = []

        if len(sheet_names) > limits.max_sheets:
            return self._failed(
                ParseStatus.REJECTED,
                ParseReasonCode.SHEET_LIMIT,
                f"工作表数量 {len(sheet_names)} 超过上限 {limits.max_sheets}",
            )

        tables: list[ParsedTable] = []
        blocks: list[TextBlock] = []
        total_rows = 0
        formula_without_cache = 0

        for sheet_name in sheet_names:
            ws_f: Worksheet = wb_formula[sheet_name]
            ws_v: Worksheet = wb_values[sheet_name]

            merged = tuple(str(rng) for rng in ws_f.merged_cells.ranges)
            merged_lookup = _merged_lookup(ws_f)

            grid, rows_scanned, sheet_warnings, no_cache = _scan_sheet(
                ws_f, ws_v, sheet_name, merged_lookup, limits
            )
            warnings.extend(sheet_warnings)
            formula_without_cache += no_cache
            total_rows += rows_scanned

            if total_rows > limits.max_total_rows:
                return self._failed(
                    ParseStatus.REJECTED,
                    ParseReasonCode.ROW_LIMIT,
                    f"数据行总数超过上限 {limits.max_total_rows}",
                )

            if not grid:
                continue

            tables.append(
                ParsedTable(
                    table_id=f"{sheet_name}",
                    sheet_name=sheet_name,
                    header_rows=(),  # 由提取层填充
                    cells=grid,
                    merged_ranges=merged,
                )
            )
            blocks.extend(_text_blocks(grid))

        if not tables:
            return self._failed(
                ParseStatus.REJECTED,
                ParseReasonCode.NO_TABLE_FOUND,
                "工作簿中没有任何非空单元格",
            )

        status = ParseStatus.NEEDS_REVIEW if warnings else ParseStatus.OK
        return ParsedDocument(
            parser_name=self.name,
            parser_version=self.version,
            status=status,
            reason_code=(ParseReasonCode.FORMULA_WITHOUT_CACHE if formula_without_cache else None),
            diagnostics={
                "sheet_count": len(sheet_names),
                "total_rows": total_rows,
                "formula_without_cache": formula_without_cache,
            },
            tables=tuple(tables),
            blocks=tuple(blocks),
            warnings=tuple(warnings),
        )


# ------------------------------------------------------------------------ helpers


def _read_magic(path: Path, size: int = 8) -> bytes:
    try:
        with path.open("rb") as fh:
            return fh.read(size)
    except OSError:
        return b""


def _check_zip_bomb(path: Path) -> ParseCapability | None:
    """解压炸弹检测。返回 None 表示通过。"""
    try:
        with zipfile.ZipFile(path) as zf:
            uncompressed = sum(info.file_size for info in zf.infolist())
            compressed = sum(info.compress_size for info in zf.infolist()) or 1
    except zipfile.BadZipFile:
        return ParseCapability.reject(ParseReasonCode.CORRUPT, "zip 结构损坏")
    except OSError:
        return ParseCapability.reject(ParseReasonCode.CORRUPT, "无法读取文件")

    if uncompressed > MAX_UNCOMPRESSED_BYTES:
        return ParseCapability.reject(
            ParseReasonCode.FILE_TOO_LARGE,
            f"解压后体积 {uncompressed // (1024 * 1024)}MB 超过上限",
        )
    if uncompressed // compressed > MAX_COMPRESSION_RATIO:
        return ParseCapability.reject(
            ParseReasonCode.FILE_TOO_LARGE,
            f"压缩比 {uncompressed // compressed}:1 异常，疑似解压炸弹",
        )
    return None


def _merged_lookup(ws: Worksheet) -> dict[tuple[int, int], str]:
    """(row, col) -> 所属合并区域字符串。用于合并大标题定位。"""
    lookup: dict[tuple[int, int], str] = {}
    for rng in ws.merged_cells.ranges:
        text = str(rng)
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                lookup[(row, col)] = text
    return lookup


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return str(value)


def _scan_sheet(
    ws_f: Worksheet,
    ws_v: Worksheet,
    sheet_name: str,
    merged_lookup: dict[tuple[int, int], str],
    limits: ParseLimits,
) -> tuple[tuple[tuple[ParsedCell, ...], ...], int, list[str], int]:
    """遍历一个 sheet，产出规整网格。

    **边遍历边数**行数，不读 ws.max_row 做预检（那个值由文件自己声明，不可信）。
    """
    rows_f = ws_f.iter_rows()
    rows_v = ws_v.iter_rows()

    grid: list[tuple[ParsedCell, ...]] = []
    warnings: list[str] = []
    formula_without_cache = 0
    rows_scanned = 0
    last_non_empty_row = -1

    for row_f, row_v in zip(rows_f, rows_v, strict=False):
        if rows_scanned >= limits.max_rows_per_sheet:
            warnings.append(
                f"工作表 {sheet_name!r} 行数超过上限 {limits.max_rows_per_sheet}，已截断"
            )
            break
        rows_scanned += 1

        parsed_row: list[ParsedCell] = []
        row_has_content = False

        for cell_f, cell_v in zip(row_f, row_v, strict=False):
            if not isinstance(cell_f, Cell):  # pragma: no cover - 合并区域的只读代理
                continue
            raw_formula = cell_f.value
            data_type = str(cell_f.data_type)
            is_formula = data_type == "f"

            cached = _cell_text(cell_v.value) if isinstance(cell_v, Cell) else None
            typed = cell_v.value if isinstance(cell_v, Cell) else None

            if is_formula and cached is None:
                formula_without_cache += 1

            value_raw = _cell_text(raw_formula)
            if is_formula:
                # 公式单元格对下游呈现的是缓存值；公式串单独留档
                value_raw = cached
                typed = cell_v.value if isinstance(cell_v, Cell) else None

            if value_raw is not None and value_raw != "":
                row_has_content = True

            ref = CellRef(
                sheet_name=sheet_name,
                row_index=cell_f.row,
                col_index=cell_f.column,
                address=f"{get_column_letter(cell_f.column)}{cell_f.row}",
                coordinate_space=CoordinateSpace.XLSX_GRID,
            )
            parsed_row.append(
                ParsedCell(
                    ref=ref,
                    value_raw=value_raw,
                    value_typed=typed,
                    data_type=data_type,
                    formula=str(raw_formula) if is_formula else None,
                    cached_value=cached if is_formula else None,
                    merged_range=merged_lookup.get((cell_f.row, cell_f.column)),
                )
            )

        grid.append(tuple(parsed_row))
        if row_has_content:
            last_non_empty_row = len(grid) - 1

    # 去掉尾部整片空行（openpyxl 常把格式化过的空行也算进来）
    trimmed = tuple(grid[: last_non_empty_row + 1]) if last_non_empty_row >= 0 else ()

    if formula_without_cache:
        warnings.append(
            f"工作表 {sheet_name!r} 有 {formula_without_cache} 个公式没有缓存值，"
            "这些单元格无法取到数值（请用 Excel 打开另存后重试）"
        )

    return trimmed, rows_scanned, warnings, formula_without_cache


def _text_blocks(grid: tuple[tuple[ParsedCell, ...], ...]) -> list[TextBlock]:
    """非空单元格 -> TextBlock。

    block_id 是 MVP-1 LLM 适配器唯一被允许引用的东西——适配器只能「选」不能「写」。
    """
    blocks: list[TextBlock] = []
    for row in grid:
        for cell in row:
            if cell.value_raw is None or cell.value_raw == "":
                continue
            blocks.append(
                TextBlock(
                    block_id=f"{cell.ref.sheet_name}!{cell.ref.address}",
                    text=cell.value_raw,
                    ref=cell.ref,
                )
            )
            if len(blocks) >= MAX_BLOCKS_PER_SHEET:
                return blocks
    return blocks
