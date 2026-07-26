"""测试公共夹具。

`make_xlsx` 用 openpyxl 现场构造工作簿——**不依赖任何受版权或隐私保护的真实客户文件**
（SPEC §16）。
"""

from __future__ import annotations

import hashlib
import sys
from collections.abc import Sequence
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

from app.parsers.base import DocumentInput

Row = Sequence[Any]


def write_xlsx(
    path: Path,
    sheets: dict[str, Sequence[Row]],
    merges: dict[str, Sequence[str]] | None = None,
) -> Path:
    """按 {工作表名: 行列表} 写出 .xlsx。"""
    wb = Workbook()
    default = wb.active
    assert default is not None
    first = True
    for name, rows in sheets.items():
        ws = default if first else wb.create_sheet()
        ws.title = name
        first = False
        for row in rows:
            ws.append(list(row))
        for rng in (merges or {}).get(name, ()):
            ws.merge_cells(rng)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def document_input(path: Path) -> DocumentInput:
    data = path.read_bytes()
    return DocumentInput(
        path=path,
        original_filename=path.name,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        file_size=len(data),
        sha256=hashlib.sha256(data).hexdigest(),
    )


@pytest.fixture
def make_xlsx(tmp_path: Path):  # type: ignore[no-untyped-def]
    """返回一个 (name, sheets, merges) -> DocumentInput 的构造函数。"""

    def _make(
        name: str,
        sheets: dict[str, Sequence[Row]],
        merges: dict[str, Sequence[str]] | None = None,
    ) -> DocumentInput:
        path = write_xlsx(tmp_path / name, sheets, merges)
        return document_input(path)

    return _make


# --------------------------------------------------------------------------------
# 三个「脏 XLSX」版面（SPEC §16.6）
# --------------------------------------------------------------------------------


def order_rows(
    *,
    title: str,
    doc_label: str,
    doc_no: str,
    date: str,
    items: Sequence[tuple[str, str, int, str, str, str]],
    grand_total: str,
    currency: str = "USD",
    incoterm: str = "FOB Ningbo, Incoterms 2020",
    payment: str = "30% T/T in advance, 70% before shipment",
    delivery: str = "30 days after receipt of deposit",
    buyer: str = "ACME TRADING GMBH",
    seller: str = "NINGBO SUNRISE IMP & EXP CO., LTD",
) -> list[Row]:
    """构造一份标准外贸单据的行数据。

    版面刻意做成真实样子：抬头 -> 标题 -> Key-Value 区 -> 行项目表 -> 合计行。
    """
    rows: list[Row] = [
        [seller],
        ["Add: No.99 Zhongshan Rd, Ningbo, China"],
        [],
        [title],
        [],
        [doc_label, doc_no, "", "Date", date],
        ["Buyer", buyer, "", "Seller", seller],
        ["Currency", currency, "", "Trade Terms", incoterm],
        ["Payment Terms", payment],
        ["Delivery Time", delivery],
        [],
        ["Item No.", "Description", "Qty", "Unit", "Unit Price", "Amount"],
    ]
    rows.extend(list(item) for item in items)
    rows.append([])
    rows.append(["", "", "", "", "Grand Total", grand_total])
    return rows


#: 标准三件套的共同行项目（fixture「三份完全一致」用）。
BASE_ITEMS: list[tuple[str, str, int, str, str, str]] = [
    ("AB-100", "Ceramic Mug 350ml", 1000, "PCS", "1.25", "1250.00"),
    ("AB-200", "Ceramic Plate 8in", 500, "PCS", "2.40", "1200.00"),
    ("AB-300", "Glass Bowl 500ml", 200, "PCS", "3.10", "620.00"),
]
BASE_GRAND_TOTAL = "3070.00"


#: A：前 5 行公司抬头 + 合并大标题
DIRTY_A: dict[str, Sequence[Row]] = {
    "Quotation": [
        ["NINGBO SUNRISE IMP & EXP CO., LTD"],
        ["Add: No.99 Zhongshan Rd, Ningbo, China"],
        ["Tel: +86-574-1234567"],
        ["Email: sales@example.com"],
        [],
        ["QUOTATION"],
        ["Quotation No.", "Q2026-001", "", "Date", "2026-07-15"],
        [],
        ["Item No.", "Description", "Qty", "Unit", "Unit Price", "Amount"],
        ["AB-100", "Ceramic Mug 350ml", 1000, "PCS", "1.25", "1250.00"],
        ["AB-200", "Ceramic Plate 8in", 500, "PCS", "2.40", "1200.00"],
        ["AB-300", "Glass Bowl 500ml", 200, "PCS", "3.10", "620.00"],
        ["", "", "", "", "Total", "3070.00"],
    ]
}
DIRTY_A_MERGES = {"Quotation": ["A1:F1", "A6:F6"]}

#: B：双行表头 + 单位行
DIRTY_B: dict[str, Sequence[Row]] = {
    "PI": [
        ["PROFORMA INVOICE"],
        [],
        ["Item", "Product", "", "Packing", "", "Price", ""],
        ["No.", "Description", "Qty", "PCS/CTN", "CTNS", "Unit Price", "Amount"],
        ["", "", "PCS", "", "", "USD", "USD"],
        ["AB-100", "Ceramic Mug 350ml", 1000, 50, 20, "1.25", "1250.00"],
        ["AB-200", "Ceramic Plate 8in", 500, 25, 20, "2.40", "1200.00"],
    ]
}
DIRTY_B_MERGES = {"PI": ["A1:G1", "D3:E3"]}

#: C：中段空行 + 尾部小计/运费/合计 + 右侧备注列（纯中文表头）
DIRTY_C: dict[str, Sequence[Row]] = {
    "订单": [
        ["型号", "品名", "数量", "单位", "单价", "金额", "备注"],
        ["AB-100", "陶瓷杯 350ml", 1000, "PCS", "1.25", "1250.00", "客户指定包装"],
        ["AB-200", "陶瓷盘 8寸", 500, "PCS", "2.40", "1200.00", ""],
        [],
        ["AB-300", "玻璃碗 500ml", 200, "PCS", "3.10", "620.00", "新品"],
        [],
        ["", "", "", "", "小计", "3070.00"],
        ["", "", "", "", "运费", "200.00"],
        ["", "", "", "", "合计", "3270.00"],
    ]
}


def pytest_sessionfinish(session: pytest.Session, exitstatus: int) -> None:
    """SPEC §17：golden 的每组指标写进 `docs/golden-report.md`。

    三道闸门，缺一份交付证据就会被悄悄改写：

    1. 只在 `tests.test_golden` 真被收集过时才动手 —— `pytest tests/test_guards.py`
       这类定向运行不该顺手覆盖它。
    2. **本轮必须全绿**（`exitstatus == 0`）。红着还写是最危险的一种：
       `_RESULTS` 是管线跑出来的，断言失败不影响它，于是
       `test_重复运行的差异顺序完全一致` 或 `test_仓库内的fixture就是生成器的产物`
       挂掉时，渲染出的报告依然是一份完美的全绿表 —— 那两条恰恰是报告本身
       看不出来的失效。跑红了就保留上一份已知良好的文件。
    3. 16 组少跑一组即拒绝落盘（在 `test_golden.write_golden_report()` 里）。
    """
    module = sys.modules.get("tests.test_golden")
    if module is None or exitstatus != 0:
        return
    module.write_golden_report()
