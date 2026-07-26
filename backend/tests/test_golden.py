"""Golden fixtures 比对。SPEC §16.2、§16.3、§16.4。

对 `fixtures/expected/*.json` 里的每一组用例跑 `app.pipeline` 全流程，断言：

  1. `required_differences` 逐条命中（按 scope+subject_key+field_name+type 匹配，
     severity 也必须相等）—— **缺一即失败**
  2. 不在 required 里的 CRITICAL 条数 <= `max_unexpected_critical`（恒为 0）
  3. 每条差异至少关联一条存在的 Evidence（Gate-0 第 9 条全量断言）
  4. 差异总数 == `total_differences`（快照值）
  5. 重复运行的 `difference_key` 列表**逐位相同**（顺序敏感，SPEC §16.4：
     禁止 `sorted(actual) == sorted(expected)` 绕过）

fixtures 缺文件时**自动补齐缺的那几个**，不 skip（硬约束 #16：pytest 输出零 skip）；
**已存在的文件一律不覆盖**，否则「仓库内产物 == 生成器产物」这条断言会被自己抹平。

断言策略的理由（SPEC §16.3）：若按「差异总数完全相等」做唯一断言，任何规则微调
都会让全部用例同时红，实施者的自然反应是改期望值——正是执行原则里明令禁止的行为。
所以硬断言只压在**植入的关键差异**与 **CRITICAL 误报上限**上，总数只作快照。

**自产 fixture 免责声明（SPEC §16.5）**：本组 fixtures 由 `tools.fixtures.build`
程序化自产，生成器与提取器共享同一套别名表与列布局假设。全绿只证明确定性比较逻辑
与解析管线的自洽与稳定，**不代表对真实客户文件的提取准确率**。
"""

from __future__ import annotations

import hashlib
import json
import tempfile
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.enums import SIGNATURE_TAGS, DocumentRole, Severity
from app.domain.models import DifferenceDraft
from app.pipeline import ProjectResult, process_document, run_project
from tests.conftest import document_input
from tools.fixtures.build import (
    FILENAMES,
    FIXTURES_ROOT,
    LAYOUT_CHINESE,
    LAYOUT_DEFAULT,
    LAYOUT_SECOND_SHEET,
    MERGED_TITLE_RANGES,
    REPO_ROOT,
    SELLER,
    TITLES,
    build_all,
    build_cases,
    case_files,
)

pytestmark = pytest.mark.golden

EXPECTED_DIR = FIXTURES_ROOT / "expected"
ORDERS_DIR = FIXTURES_ROOT / "orders"

#: 差异的匹配签名。**刻意不含具体数值与文案**——golden 不得对 explanation 文本
#: 做字符串比对，否则措辞微调红一片（SPEC §13.1）。
Signature = tuple[str, str, str, str]


def _ensure_fixtures() -> None:
    """补齐缺失的 fixture 文件。**不 skip**（硬约束 #16）。

    **只补缺的，绝不覆盖已存在的。** 直接调 `build_all(FIXTURES_ROOT)` 会把仓库里
    已有的 `expected/*.json` 一并重写：那样 `test_仓库内的fixture就是生成器的产物`
    永远比较「刚写出来的」和「刚写出来的」，手改过的期望值会被这一步悄悄抹平，
    §16.5 免责声明所依赖的「fixture 确由程序自产」也就无从验证。

    要重新生成期望值请显式跑 `python -m tools.fixtures.build` —— 那是一次
    有意的、会进 git diff 的操作，而不是跑测试的副作用。
    """
    missing = [path for case in build_cases() for path in case_files(case) if not path.exists()]
    if not missing:
        return
    with tempfile.TemporaryDirectory() as workdir:
        staging = Path(workdir)
        build_all(staging)
        for path in missing:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes((staging / path.relative_to(FIXTURES_ROOT)).read_bytes())


_ensure_fixtures()

CASE_IDS: tuple[str, ...] = tuple(sorted(p.stem for p in EXPECTED_DIR.glob("*.json")))

#: 同一组用例被多个测试函数复用，跑一次缓存起来（纯函数，缓存不会掩盖问题）。
_RESULTS: dict[str, ProjectResult] = {}


def _load_expected(case_id: str) -> dict[str, Any]:
    payload = json.loads((EXPECTED_DIR / f"{case_id}.json").read_text(encoding="utf-8"))
    assert isinstance(payload, dict)
    return payload


def _run_case(case_id: str, expected: dict[str, Any]) -> ProjectResult:
    """按 expected.json 声明的角色集合跑完整管线（与 API 层同一条路径）。"""
    processed = {}
    for role_value in expected["compared_roles"]:
        role = DocumentRole(role_value)
        path: Path = ORDERS_DIR / case_id / FILENAMES[role]
        processed[role] = process_document(
            document_id=f"{case_id}:{role.value.lower()}",
            role=role,
            src=document_input(path),
        )
    return run_project(f"golden:{case_id}", processed)


def _result(case_id: str) -> ProjectResult:
    if case_id not in _RESULTS:
        _RESULTS[case_id] = _run_case(case_id, _load_expected(case_id))
    return _RESULTS[case_id]


def _signature_of(diff: DifferenceDraft) -> Signature:
    return (
        diff.scope.value,
        diff.subject_key,
        diff.field_name or "",
        diff.difference_type.value,
    )


def _expected_signature(item: dict[str, Any]) -> Signature:
    return (
        str(item["scope"]),
        str(item["subject_key"]),
        str(item["field_name"] or ""),
        str(item["difference_type"]),
    )


def _describe(diff: DifferenceDraft) -> str:
    return (
        f"{diff.scope.value}/{diff.subject_key}/{diff.field_name or '-'}"
        f"/{diff.difference_type.value}/{diff.severity.value}"
    )


def test_用例集合非空且已全部生成() -> None:
    """fixtures 缺失时生成器必须已经补齐——不允许靠 skip 蒙混过关。"""
    planned = tuple(sorted(case.case_id for case in build_cases()))
    assert planned == CASE_IDS, "fixtures/expected 与生成器声明的用例不一致"
    assert len(CASE_IDS) >= 16, (
        "MVP-0 要求 12 组语义（SPEC §16.2）+ 1 组两文件变体 + 3 组版面变体（SPEC §16.1），"
        f"实际 {len(CASE_IDS)} 组"
    )


@pytest.mark.parametrize("case_id", CASE_IDS)
def test_参与比较的文档全部解析可用(case_id: str) -> None:
    """先证明「零差异」不是因为解析失败。

    没有这一条，一个把所有文档判成 REJECTED 的回归会让「三份完全一致」组
    以 0 条差异通过——最危险的那种绿。
    """
    expected = _load_expected(case_id)
    result = _result(case_id)
    declared = {DocumentRole(value) for value in expected["compared_roles"]}
    for role, processing in result.processing.items():
        assert processing.usable, (
            f"{case_id} 的 {role.value} 未能产出快照："
            f"{processing.parsed.status.value}/{processing.parsed.reason_code}"
        )
        assert processing.snapshot is not None
        assert processing.snapshot.line_items, f"{case_id} 的 {role.value} 没有提取到行项目"
    assert result.snapshot.compared_roles == frozenset(declared)
    assert result.snapshot.runnable


#: 生成器**每一份**单据都写了这些文档级字段（见 `build.document_rows`）。
#: 少一个就说明「标签 -> 字段」这条链路退化了，而不是「这类单据本来就不写」。
REQUIRED_DOC_FIELDS: tuple[str, ...] = (
    "document_number",
    "document_date",
    "buyer_name",
    "seller_name",
    "currency",
    "incoterm",
    "incoterm_named_place",
    "incoterm_version",
    "payment_terms",
    "delivery_terms",
    "grand_total",
)

#: 生成器**每一行**都写了这些行项目字段（六列版面）。
REQUIRED_LINE_FIELDS: tuple[str, ...] = (
    "internal_sku",
    "description",
    "quantity",
    "unit",
    "unit_price",
    "line_total",
)


@pytest.mark.parametrize("case_id", CASE_IDS)
def test_每份单据的字段确实被提取到(case_id: str) -> None:
    """堵住最危险的一种绿：**什么都没提取到，也是零差异**。

    `identical` / `row_order_shuffled` / `two_docs_only` / `layout_merged_title`
    期望 0 条差异。一个让文档级提取整体失效的回归（别名表被改、Key-Value 扫描
    被改坏、中文标签不再命中）不会让它们变红——全没提取到的字段在 SPEC §9.8 下
    本来就不产出 MISSING_VALUE，差异总数照样是 0，required 集合为空也照样通过。
    Gate-0 第 5 条那条最重要的断言会在无人察觉时变成一句空话。

    所以必须正面断言：生成器写进去的字段，管线确实读出来了。
    """
    result = _result(case_id)
    problems: list[str] = []
    for role in sorted(result.snapshot.documents, key=lambda r: r.value):
        document = result.snapshot.documents[role]
        for key in REQUIRED_DOC_FIELDS:
            if not document.get(key).present:
                problems.append(f"{role.value} 缺文档级字段 {key}")
        for item in document.line_items:
            for key in REQUIRED_LINE_FIELDS:
                if not item.get(key).present:
                    problems.append(f"{role.value} 的 {item.line_key} 缺行项目字段 {key}")
    assert not problems, (
        f"{case_id} 的提取结果不完整（生成器写了但管线没读到）：{problems}。"
        "零差异组一旦提取失效仍会全绿，所以这条断言不能省。"
    )


@pytest.mark.parametrize("case_id", CASE_IDS)
def test_必须命中的差异一条都不能少(case_id: str) -> None:
    """硬断言：植入的关键差异全部被发现（Gate-0 第 4 条，召回 100%）。"""
    expected = _load_expected(case_id)
    actual = {_signature_of(d): d for d in _result(case_id).comparison.differences}

    for item in expected["required_differences"]:
        signature = _expected_signature(item)
        assert signature in actual, (
            f"{case_id} 缺少必须命中的差异 {signature}；"
            f"实际差异：{[_describe(d) for d in _result(case_id).comparison.differences]}"
        )
        hit = actual[signature]
        assert hit.severity.value == item["severity"], (
            f"{case_id} 的 {signature} 风险等级应为 {item['severity']}，实际 {hit.severity.value}"
        )


@pytest.mark.parametrize("case_id", CASE_IDS)
def test_非植入的critical误报不超过上限(case_id: str) -> None:
    """Gate-0 第 5 条：`max_unexpected_critical` 恒为 0，一条误报都不许有。"""
    expected = _load_expected(case_id)
    planted = {_expected_signature(item) for item in expected["required_differences"]}
    unexpected = [
        d
        for d in _result(case_id).comparison.differences
        if d.severity is Severity.CRITICAL and _signature_of(d) not in planted
    ]
    limit = int(expected["max_unexpected_critical"])
    assert limit == 0, "max_unexpected_critical 必须恒为 0（Gate-0 第 5 条）"
    assert len(unexpected) <= limit, (
        f"{case_id} 出现 {len(unexpected)} 条非植入的 CRITICAL 误报："
        f"{[_describe(d) for d in unexpected]}"
    )


@pytest.mark.parametrize("case_id", CASE_IDS)
def test_每条差异都至少有一条证据(case_id: str) -> None:
    """Gate-0 第 9 条全量断言：没有证据的差异等于没法核对。"""
    result = _result(case_id)
    evidence = result.evidence
    for diff in result.comparison.differences:
        assert diff.evidence_ids, f"{case_id} 的 {_describe(diff)} 没有关联任何证据"
        for evidence_id in diff.evidence_ids:
            assert evidence_id in evidence, (
                f"{case_id} 的 {_describe(diff)} 引用了不存在的证据 {evidence_id}"
            )


@pytest.mark.parametrize("case_id", CASE_IDS)
def test_差异总数与快照一致(case_id: str) -> None:
    """快照断言（SPEC §16.3）。"""
    expected = _load_expected(case_id)
    differences = _result(case_id).comparison.differences
    assert len(differences) == int(expected["total_differences"]), (
        f"{case_id} 的差异总数为 {len(differences)}，"
        f"expected.json 记的是 {expected['total_differences']}。\n"
        "这是**快照值**：先确认规则变更是否合理（是不是漏报/误报），"
        "确认合理后显式更新 fixtures/expected/"
        f"{case_id}.json 并在 commit message 说明理由；"
        "**不得为了让测试变绿而改期望**。\n"
        f"实际差异：{[_describe(d) for d in differences]}"
    )


@pytest.mark.parametrize("case_id", CASE_IDS)
def test_重复运行的差异顺序完全一致(case_id: str) -> None:
    """Gate-0 第 15 条 + SPEC §16.4：**顺序敏感**比较，禁止排序后再比。"""
    expected = _load_expected(case_id)
    first = [d.difference_key for d in _result(case_id).comparison.differences]
    second = [d.difference_key for d in _run_case(case_id, expected).comparison.differences]
    assert first == second, f"{case_id} 两次运行的差异顺序或内容不一致"

    differences = _result(case_id).comparison.differences
    sort_keys = [d.sort_key() for d in differences]
    assert sort_keys == sorted(sort_keys), (
        f"{case_id} 的差异未按 (scope, subject_key, field_name, difference_type) 稳定排序"
    )
    assert len(set(first)) == len(first), f"{case_id} 出现重复的 difference_key"


@pytest.mark.parametrize("case_id", CASE_IDS)
def test_摘要与前提摘要稳定(case_id: str) -> None:
    """values_digest 必须由取值决定，重跑同一输入不得漂移。

    它是「前提是否变化」的唯一判据（SPEC §11.3），漂移会让审核状态
    在用户什么都没改的情况下集体置 NEEDS_CONFIRMATION。
    """
    expected = _load_expected(case_id)
    first = {d.difference_key: d.values_digest for d in _result(case_id).comparison.differences}
    second = {
        d.difference_key: d.values_digest
        for d in _run_case(case_id, expected).comparison.differences
    }
    assert first == second


def _digests(root: Path) -> dict[str, str]:
    return {
        path.relative_to(root).as_posix(): hashlib.sha256(path.read_bytes()).hexdigest()
        for path in build_all(root)
    }


def test_重复生成的fixture逐字节一致(tmp_path: Path) -> None:
    """SPEC §16.1：不用随机数，二次生成 sha256 必须一致。

    xlsx 是 zip，openpyxl 会写入当前时间戳；生成器按固定时间戳重打包，
    因此这里直接比 sha256，而不是退让成「只比解析结果」。
    """
    first = _digests(tmp_path / "run1")
    second = _digests(tmp_path / "run2")
    assert first, "生成器没有产出任何文件"
    assert first == second, "同一组参数两次生成的产物不一致（存在非确定性来源）"


def _sheet_rows(case_id: str, role: DocumentRole) -> tuple[Worksheet, list[list[object]]]:
    """读回生成好的 xlsx，返回订单表所在工作表与它的网格。"""
    workbook = load_workbook(ORDERS_DIR / case_id / FILENAMES[role], data_only=False)
    sheet = workbook[workbook.sheetnames[-1]]
    return sheet, [list(row) for row in sheet.iter_rows(values_only=True)]


def test_版面变体确实改变了版面() -> None:
    """三组版面变体复用了别组的期望值 —— **前提是它们的版面真的不一样**。

    变体一旦悄悄退化回默认版面（`Layout` 字段写错、`workbook_bytes` 忘了用
    `layout`），三组就成了三份重复副本：照样全绿，却一寸新东西都没测，
    而「换版面差异不变」这个结论也就没有任何证据支撑。
    """
    role = DocumentRole.QUOTATION

    # 变体一：公司抬头与大标题被合并成整行；默认版面不得有任何合并区（对照组）。
    merged_sheet, merged_rows = _sheet_rows("layout_merged_title", role)
    merged = {str(rng) for rng in merged_sheet.merged_cells.ranges}
    assert merged == set(MERGED_TITLE_RANGES), f"合并区实际为 {sorted(merged)}"
    assert merged_rows[0][0] == SELLER, "第 1 行不再是公司抬头，MERGED_TITLE_RANGES 已错位"
    assert merged_rows[3][0] == TITLES[role], "第 4 行不再是大标题，MERGED_TITLE_RANGES 已错位"

    baseline_sheet, _ = _sheet_rows("identical", role)
    assert not baseline_sheet.merged_cells.ranges, "默认版面不该有合并区，对照失去意义"

    # 变体二：标签与列头全是中文。
    _, chinese_rows = _sheet_rows("layout_chinese_headers", role)
    assert tuple(chinese_rows[11][:6]) == LAYOUT_CHINESE.headers
    assert chinese_rows[8][0] == LAYOUT_CHINESE.payment_terms  # 必命中差异所依赖的标签
    assert LAYOUT_CHINESE.headers != LAYOUT_DEFAULT.headers, "中文版面与默认版面必须不同"

    # 变体三：订单表在第二个 sheet，第一个 sheet 过不了表头门槛。
    workbook = load_workbook(ORDERS_DIR / "layout_second_sheet" / FILENAMES[role], data_only=False)
    assert workbook.sheetnames == [LAYOUT_SECOND_SHEET.lead_sheet_title, TITLES[role]]
    single = load_workbook(ORDERS_DIR / "identical" / FILENAMES[role], data_only=False)
    assert len(single.sheetnames) == 1, "默认版面必须是单 sheet，否则对照失去意义"


# --------------------------------------------------------------------------------
# SPEC §17：每组一张表，写进 docs/golden-report.md
# --------------------------------------------------------------------------------

#: 最终交付报告的「测试数据结果」一节必须是本文件的原样粘贴（SPEC §17）。
REPORT_PATH = REPO_ROOT / "docs" / "golden-report.md"

#: §16.5 免责声明。这份报告会被整段贴进交付报告，"16 组全绿"若不带着这句话出现，
#: 读者会把它读成"对真实客户文件的准确率"——而 fixtures 与提取器共享同一套别名表。
DISCLAIMER = (
    "本组 fixtures 由 `tools.fixtures.build` 程序化自产，生成器与提取器共享同一套"
    "别名表与列布局假设。全绿只证明确定性比较逻辑与解析管线的自洽与稳定，"
    "**不代表对真实客户文件的提取准确率**。"
)


#: 一组用例的一行明细：(差异, 它是否为 expected.json 里植入的必命中差异)。
ReportRow = tuple[DifferenceDraft, bool]

#: 角色缩写，与 `role_signature` 共用 `SIGNATURE_TAGS`（唯一真源，SPEC §3.2）。
_ROLE_TAG: dict[DocumentRole, str] = {role: tag for tag, role in SIGNATURE_TAGS}


def _report_rows(case_id: str) -> tuple[dict[str, Any], list[ReportRow]]:
    """返回 (expected, [(差异, 是否为植入的必命中差异)])。"""
    expected = _load_expected(case_id)
    planted = {_expected_signature(item) for item in expected["required_differences"]}
    rows = [(d, _signature_of(d) in planted) for d in _result(case_id).comparison.differences]
    return expected, rows


def _overview_row(case_id: str, expected: dict[str, Any], rows: list[ReportRow]) -> str:
    """总览表的一行。**单独成函数是为了能被正面测到**——「非植入 CRITICAL」
    这一格是 Gate-0 第 5 条的门面，它要是恒为 0，报告在真出误报时也照样显示 0。"""
    # 不能取首字母：PURCHASE_ORDER 与 PROFORMA_INVOICE 都是 P，两文件组会显示成
    # 「P+P」，读者无从判断缺席的是报价单还是采购订单。走和 role_signature 同一张表。
    roles = "+".join(_ROLE_TAG[DocumentRole(r)] for r in expected["compared_roles"])
    required = len(expected["required_differences"])
    hit = sum(1 for _, planted in rows if planted)
    unexpected_critical = sum(
        1 for diff, planted in rows if diff.severity is Severity.CRITICAL and not planted
    )
    return f"| `{case_id}` | {roles} | {required} | {hit} | {unexpected_critical} | {len(rows)} |"


def render_golden_report() -> str:
    """把已跑过的用例渲染成 Markdown。**顺序即 golden 的输出顺序，不重排。**

    刻意不写生成时间：这份文件的价值在于"同样的代码永远产出同样的表"，
    塞一个每次都变的时间戳会让它每跑一次就产生一次假 diff，
    真正的指标变化反而淹没在噪声里。时间戳由 `validation-report.md` 那边记。
    """
    lines = [
        "# Golden 用例指标报告",
        "",
        "> **本文件由 `pytest` 的 `pytest_sessionfinish` 钩子自动生成，禁止手写。**",
        f"> 覆盖 {len(CASE_IDS)} 组 fixtures（SPEC §16.2 的 12 组语义 + 1 组两文件变体"
        " + 3 组版面变体）。",
        "",
        f"> ⚠️ **自产 fixture 免责声明（SPEC §16.5）**：{DISCLAIMER}",
        "",
        "## 总览",
        "",
        "| 用例 | 参与角色 | 必须命中 | 实际命中 | 非植入 CRITICAL | 差异总数 |",
        "|---|---|---:|---:|---:|---:|",
    ]

    detail: list[str] = []
    total_required = total_hit = total_unexpected = 0
    for case_id in CASE_IDS:
        expected, rows = _report_rows(case_id)
        lines.append(_overview_row(case_id, expected, rows))
        total_required += len(expected["required_differences"])
        total_hit += sum(1 for _, planted in rows if planted)
        total_unexpected += sum(
            1 for d, planted in rows if d.severity is Severity.CRITICAL and not planted
        )

        detail += [
            "",
            f"### `{case_id}`",
            "",
            f"{expected['description']}",
            "",
        ]
        if not rows:
            detail.append("本组期望零差异，实际产出 **0 条**。")
            continue
        detail += [
            "| 范围 | 主体 | 字段 | 差异类型 | 风险等级 | 植入 |",
            "|---|---|---|---|---|---|",
        ]
        detail += [
            f"| {d.scope.value} | `{d.subject_key}` | {d.field_name or '—'} "
            f"| {d.difference_type.value} | {d.severity.value} | {'✅' if planted else '—'} |"
            for d, planted in rows
        ]

    # 召回率与误报数都从实际产出里数出来。写成 `{total_required}/{total_required}`
    # 是句同义反复：漏报时硬断言会红，但这份被当作证据粘贴的文件仍会显示 100%。
    lines += [
        "",
        f"**合计**：{len(CASE_IDS)} 组，植入差异 {total_required} 条，"
        f"实际命中 {total_hit} 条（召回 {total_hit}/{total_required}）；"
        f"非植入 CRITICAL 误报 **{total_unexpected}** 条"
        "（上限恒为 0，见 Gate-0 第 5 条）。",
        "",
        "---",
        "",
        "## 逐组明细",
        *detail,
        "",
    ]
    return "\n".join(lines)


def write_golden_report() -> Path | None:
    """全部用例都跑过时才落盘；否则原样保留旧文件并返回 None。

    `pytest -k pi_unit_price` 只跑一组也会触发 sessionfinish。不设这道闸门，
    这份被当作交付证据原样粘贴的文件就会被一次随手的定向调试覆盖成"1 组全绿"——
    而它长得和完整报告一模一样，没有任何地方会变红。
    """
    if not CASE_IDS or any(case_id not in _RESULTS for case_id in CASE_IDS):
        return None
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(render_golden_report(), encoding="utf-8", newline="\n")
    return REPORT_PATH


def test_指标报告覆盖全部用例() -> None:
    """渲染器本身要有测试——它产出的数字会被原样贴进交付报告。"""
    text = render_golden_report()
    for case_id in CASE_IDS:
        assert f"`{case_id}`" in text, f"指标报告漏了 {case_id}"
        assert _load_expected(case_id)["description"] in text
    assert DISCLAIMER in text, "§16.5 免责声明必须逐字出现（硬约束 #14）"
    assert "http://" not in text and "https://" not in text

    # 两文件组必须能看出缺席的是谁：采购订单与形式发票的首字母都是 P。
    row = next(line for line in text.splitlines() if line.startswith("| `two_docs_only` |"))
    assert row.split("|")[2].strip() == "P+I", row
    full = next(line for line in text.splitlines() if line.startswith("| `identical` |"))
    assert full.split("|")[2].strip() == "Q+P+I", full


def test_总览里的非植入critical会被真的数出来() -> None:
    """这一格是 Gate-0 第 5 条的门面：写死成 0 的话，真出误报时报告照样显示 0。

    直接拿一条真实的 CRITICAL 差异、把「植入」标成 False 喂进去 —— 走的正是
    渲染器统计的同一个函数。
    """
    expected, rows = _report_rows("po_quantity_changed")
    critical = next(d for d, _ in rows if d.severity is Severity.CRITICAL)

    clean = _overview_row("demo", expected, rows)
    assert clean.split("|")[5].strip() == "0", clean

    polluted = _overview_row("demo", expected, [*rows, (critical, False)])
    assert polluted.split("|")[5].strip() == "1", polluted


def test_报告只在全部用例跑过之后才落盘() -> None:
    """闸门本身要会咬人：少一组就不许覆盖已有的交付证据。"""
    render_golden_report()  # 先把 _RESULTS 填满，否则「返回 None」可能是别的原因
    missing = CASE_IDS[0]
    stashed = _RESULTS.pop(missing)
    try:
        assert write_golden_report() is None, "少跑了一组仍然落盘了，交付证据会被悄悄削短"
    finally:
        _RESULTS[missing] = stashed
    assert write_golden_report() == REPORT_PATH


def test_仓库内的fixture就是生成器的产物(tmp_path: Path) -> None:
    """手改过 xlsx 或 expected.json 会在这里被抓到。

    手改的 fixture 一旦被当成「程序化自产」，SPEC §16.5 的免责声明就变成了假话。
    """
    build_all(tmp_path)
    for case in build_cases():
        for path in case_files(case, tmp_path):
            committed = FIXTURES_ROOT / path.relative_to(tmp_path)
            assert committed.exists(), f"仓库内缺少 {committed.name}，请运行生成器"
            assert committed.read_bytes() == path.read_bytes(), (
                f"{committed.name} 与生成器产物不一致，请重新运行 python -m tools.fixtures.build"
            )
