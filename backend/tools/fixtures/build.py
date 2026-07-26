"""golden fixtures 生成器。SPEC §16.1、§16.2、§16.3。

产出 **12 组语义用例 + 1 组两文件（PO+PI）变体 + 3 组版面变体 = 16 组**。
版面变体（抬头+合并大标题 / 纯中文表头 / 表在第二个 sheet）按 SPEC §16.1
**复用其他组的 expected**：换版面不换语义，差异集合必须逐条不变。

产物写到**仓库根**（不在 backend/ 内）：

    fixtures/orders/<case_id>/{quotation,purchase_order,proforma_invoice}.xlsx
    fixtures/expected/<case_id>.json

用法（在 backend/ 下）：

    python -m tools.fixtures.build           生成全部用例
    python -m tools.fixtures.build --list    只列出用例，不写文件

三条不可动摇的设计约束：

1. **不用随机数。** 每一个字节都由本模块里的参数决定，因此「固定种子」这件事
   在这里退化成「根本没有种子」。重复生成必得同一份文件（SPEC §16.1 要求
   二次生成 sha256 一致）。openpyxl 保存的 zip 默认写入**当前时间戳**，会破坏
   字节级一致；因此保存后按固定时间戳重新打包一次——这是可复现构建的常规做法
   （与 strip-nondeterminism 同一手法），不是绕过校验的 hack。

2. **expected.json 由「植入差异的意图」逐条手工推导写死，绝不由跑一遍程序倒推。**
   倒推等于用实现验证实现，是循环论证：实现错了，期望跟着错，golden 照样全绿。
   每一组的 `total_differences` 下面都写清了它由哪几条差异加总而来。

3. **golden 只用英文枚举标识符**（硬约束 #1）。因此 required_differences 里的
   scope / difference_type / severity 全部取自 `app.domain.enums`，
   中文只出现在 `description` 里。

`max_unexpected_critical` **恒为 0**：非植入的 CRITICAL 就是误报，
一条都不允许（Gate-0 第 5 条）。`total_differences` 是快照值，允许显式更新，
但更新必须在 commit message 里说明理由（SPEC §16.3）。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, replace
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from app.domain.enums import DifferenceType, DocumentRole, Scope, Severity

# --------------------------------------------------------------------------------
# 路径
# --------------------------------------------------------------------------------

#: tools/fixtures/build.py -> tools -> backend -> 仓库根
REPO_ROOT = Path(__file__).resolve().parents[3]
FIXTURES_ROOT = REPO_ROOT / "fixtures"

#: 每个角色的固定文件名。golden 测试按它拼路径，改名即破坏所有已生成用例。
FILENAMES: dict[DocumentRole, str] = {
    DocumentRole.QUOTATION: "quotation.xlsx",
    DocumentRole.PURCHASE_ORDER: "purchase_order.xlsx",
    DocumentRole.PROFORMA_INVOICE: "proforma_invoice.xlsx",
}

#: 贸易链条顺序。文档、compared_roles 都按它排，绝不依赖 dict 迭代顺序。
CHAIN_ORDER: tuple[DocumentRole, ...] = (
    DocumentRole.QUOTATION,
    DocumentRole.PURCHASE_ORDER,
    DocumentRole.PROFORMA_INVOICE,
)

TITLES: dict[DocumentRole, str] = {
    DocumentRole.QUOTATION: "QUOTATION",
    DocumentRole.PURCHASE_ORDER: "PURCHASE ORDER",
    DocumentRole.PROFORMA_INVOICE: "PROFORMA INVOICE",
}

# --------------------------------------------------------------------------------
# 基准单据的内容
# --------------------------------------------------------------------------------

SELLER = "NINGBO SUNRISE IMP & EXP CO., LTD"
SELLER_ADDRESS = "Add: No.99 Zhongshan Rd, Ningbo, China"
BUYER = "ACME TRADING GMBH"

#: 三份单据引用**同一个**业务编号（真实外贸里报价/PO/PI 共用一个 Ref No. 很常见）。
#: 这样「三份完全一致」组的期望差异数就是干净的 0，Gate-0 第 5 条的断言最硬。
REFERENCE_NO = "SC-2026-0088"
DOC_DATE = "2026-07-15"

CURRENCY_USD = "USD"
CURRENCY_EUR = "EUR"

INCOTERM_FOB_NINGBO = "FOB Ningbo, Incoterms 2020"
INCOTERM_CIF_NINGBO = "CIF Ningbo, Incoterms 2020"
INCOTERM_FOB_SHANGHAI = "FOB Shanghai, Incoterms 2020"

PAYMENT_30_70 = "30% T/T in advance, 70% before shipment"
PAYMENT_50_50 = "50% T/T in advance, 50% before shipment"

DELIVERY = "30 days after receipt of deposit"

#: 打包时间戳固定成 1980-01-01（zip 格式能表达的最小时间），去掉最后一处非确定性。
_ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)
#: 文档属性里的创建/修改时间同样固定，否则 docProps/core.xml 每次都不同。
#: OOXML 的 dcterms 时间不带时区，这里刻意用 naive datetime。
_DOC_TIMESTAMP = datetime(2026, 1, 1, 0, 0, 0)
_DOC_TIMESTAMP_XML = b"2026-01-01T00:00:00Z"

#: openpyxl 的 `save_workbook()` 在写盘那一刻把 `properties.modified` 覆盖成
#: `utcnow()`，**先设好属性也没用**。这是 xlsx 里最后一处非确定性来源，
#: 打包时按固定值改回去。
#:
#: 替换**必须显式校验命中**：`re.sub` 匹配不上时会原样返回，于是时间戳静默退回
#: `utcnow()`。而 OOXML 的时间戳只精确到秒，「连跑两次比 sha256」这类测试通常在
#: 同一秒内跑完，会照常全绿——非确定性要到第二天才暴露。所以这里宁可当场抛，
#: 不接受「测试大概率会红」。
_MODIFIED_XML = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")
#: core.xml 里所有 dcterms 日期元素，用于兜底校验：openpyxl 将来新增任何一个
#: 带时间戳的元素，都会在这里被抓住，而不是变成一处新的非确定性来源。
_ANY_DCTERMS_DATE = re.compile(rb"<dcterms:(created|modified)[^>]*>([^<]*)</dcterms:\1>")
_CORE_PROPERTIES_ENTRY = "docProps/core.xml"


@dataclass(frozen=True)
class LineItem:
    """一行行项目。金额一律写成字符串，**生成器里也不出现二进制浮点**。"""

    sku: str
    description: str
    quantity: int
    unit: str
    unit_price: str
    line_total: str


MUG = LineItem("AB-100", "Ceramic Mug 350ml", 1000, "PCS", "1.25", "1250.00")
PLATE = LineItem("AB-200", "Ceramic Plate 8in", 500, "PCS", "2.40", "1200.00")
BOWL = LineItem("AB-300", "Glass Bowl 500ml", 200, "PCS", "3.10", "620.00")
TRAY = LineItem("AB-400", "Bamboo Tray 30cm", 300, "PCS", "4.00", "1200.00")

BASE_ITEMS: tuple[LineItem, ...] = (MUG, PLATE, BOWL)
#: 1250.00 + 1200.00 + 620.00
BASE_GRAND_TOTAL = "3070.00"


# --------------------------------------------------------------------------------
# 版面（SPEC §16.1 的 3 组版面变体只改这里，**一个语义字段都不动**）
# --------------------------------------------------------------------------------


@dataclass(frozen=True)
class Layout:
    """一份单据长什么样。标签与表头文本、是否合并大标题、订单表在第几个 sheet。

    版面与语义**完全正交**：同一组数据换个版面，差异集合必须逐条不变。
    这正是版面变体组「复用其他组 expected」的合法性来源（SPEC §16.1）。
    """

    layout_id: str
    #: Key-Value 区的九个标签。必须都在 FieldSpec 的**文档级**别名表里。
    ref_no: str
    date: str
    buyer: str
    seller: str
    currency: str
    trade_terms: str
    payment_terms: str
    delivery: str
    grand_total: str
    #: 行项目表的六个列头。必须都在**行项目级**别名表里。
    headers: tuple[str, str, str, str, str, str]
    #: 公司抬头与大标题是否合并成整行（真实单据极常见，且会干扰表头打分器）。
    merge_title: bool = False
    #: 非 None 时在订单表**之前**插一张说明表 —— 订单表落到第二个 sheet。
    lead_sheet_title: str | None = None


LAYOUT_DEFAULT = Layout(
    layout_id="default",
    ref_no="Ref No.",
    date="Date",
    buyer="Buyer",
    seller="Seller",
    currency="Currency",
    trade_terms="Trade Terms",
    payment_terms="Payment Terms",
    delivery="Delivery Time",
    grand_total="Grand Total",
    headers=("Item No.", "Description", "Qty", "Unit", "Unit Price", "Amount"),
)

#: 变体一：公司抬头 + 合并大标题。
LAYOUT_MERGED_TITLE = replace(LAYOUT_DEFAULT, layout_id="merged_title", merge_title=True)

#: 变体二：纯中文标签与表头。内容（金额、条款正文）保持不变——换的是版面不是语义。
#: 这是**唯一**端到端验证中文别名表的路径：解析器单测只查表头列映射，
#: 覆盖不到「中文标签 -> 文档级字段」这一段。
LAYOUT_CHINESE = Layout(
    layout_id="chinese",
    ref_no="单号",
    date="日期",
    buyer="买方",
    seller="卖方",
    currency="币种",
    trade_terms="贸易术语",
    payment_terms="付款条件",
    delivery="交货期",
    grand_total="合计",
    headers=("型号", "品名", "数量", "单位", "单价", "金额"),
)

#: 变体三：订单表在第二个 sheet。
LAYOUT_SECOND_SHEET = replace(LAYOUT_DEFAULT, layout_id="second_sheet", lead_sheet_title="附页")

#: 前置说明表的内容。**刻意不含任何别名**：命中了就会污染文档级字段提取，
#: 那样这一组测的就不再是「表在第二个 sheet」了。
LEAD_SHEET_ROWS: tuple[tuple[str, ...], ...] = (
    ("随附文件清单",),
    (),
    ("1. 产品图片",),
    ("2. 包装示意图",),
)

#: `merge_title` 版面里被合并成整行的两行（1 基行号，与 `document_rows` 一一对应）：
#: 第 1 行公司抬头、第 4 行单据大标题。改 `document_rows` 的行序必须同步改这里——
#: `test_版面变体确实改变了版面` 会在错位时立刻红。
MERGED_TITLE_RANGES: tuple[str, ...] = ("A1:F1", "A4:F4")


# --------------------------------------------------------------------------------
# 计划对象
# --------------------------------------------------------------------------------


@dataclass(frozen=True)
class DocumentPlan:
    """一份单据的全部可变内容。"""

    role: DocumentRole
    items: tuple[LineItem, ...] = BASE_ITEMS
    grand_total: str = BASE_GRAND_TOTAL
    currency: str = CURRENCY_USD
    incoterm: str = INCOTERM_FOB_NINGBO
    payment_terms: str = PAYMENT_30_70
    delivery_terms: str = DELIVERY
    layout: Layout = LAYOUT_DEFAULT

    @property
    def title(self) -> str:
        return TITLES[self.role]

    @property
    def filename(self) -> str:
        return FILENAMES[self.role]


@dataclass(frozen=True)
class DocOverride:
    """相对基准单据的改动。None 表示该字段不动。"""

    items: tuple[LineItem, ...] | None = None
    grand_total: str | None = None
    currency: str | None = None
    incoterm: str | None = None
    payment_terms: str | None = None


@dataclass(frozen=True)
class ExpectedDifference:
    """一条**必须命中**的差异。golden 按 scope+subject_key+field_name+type 匹配，
    severity 也必须相等（SPEC §16.3 的硬断言部分）。"""

    scope: Scope
    subject_key: str
    field_name: str | None
    difference_type: DifferenceType
    severity: Severity

    def to_json(self) -> dict[str, Any]:
        return {
            "scope": self.scope.value,
            "subject_key": self.subject_key,
            "field_name": self.field_name,
            "difference_type": self.difference_type.value,
            "severity": self.severity.value,
        }


@dataclass(frozen=True)
class CasePlan:
    case_id: str
    description: str
    documents: tuple[DocumentPlan, ...]
    required_differences: tuple[ExpectedDifference, ...]
    #: 快照值：本组期望产出的差异总条数（含 required 与其余低等级差异）。
    total_differences: int

    @property
    def compared_roles(self) -> tuple[DocumentRole, ...]:
        return tuple(plan.role for plan in self.documents)

    def to_json(self) -> dict[str, Any]:
        return {
            "case_id": self.case_id,
            "description": self.description,
            "compared_roles": [role.value for role in self.compared_roles],
            "required_differences": [d.to_json() for d in self.required_differences],
            # 非植入的 CRITICAL 误报上限，**恒为 0**（Gate-0 第 5 条）
            "max_unexpected_critical": 0,
            "total_differences": self.total_differences,
        }


# --------------------------------------------------------------------------------
# 期望差异的构造助手（只是省字，不含任何推断）
# --------------------------------------------------------------------------------


def _doc_diff(
    field_name: str,
    difference_type: DifferenceType,
    severity: Severity,
) -> ExpectedDifference:
    """文档级差异。subject_key 恒为 'PROJECT'（比较引擎的写法）。"""
    return ExpectedDifference(Scope.DOCUMENT, "PROJECT", field_name, difference_type, severity)


def _line_diff(
    sku: str,
    field_name: str | None,
    difference_type: DifferenceType,
    severity: Severity,
) -> ExpectedDifference:
    """行项目差异。subject_key 是匹配组的 group_key = 'SKU:<归一SKU>'。"""
    return ExpectedDifference(Scope.LINE_ITEM, f"SKU:{sku}", field_name, difference_type, severity)


def _calc_line_diff(role: DocumentRole, sku: str, severity: Severity) -> ExpectedDifference:
    """文档内行金额校验。subject_key = '<ROLE>#<line_key>'，line_key 见 §3.3。"""
    return ExpectedDifference(
        Scope.CALCULATION,
        f"{role.value}#sku:{sku}#1",
        "line_total",
        DifferenceType.CALCULATION_ERROR,
        severity,
    )


def _calc_total_diff(role: DocumentRole, severity: Severity) -> ExpectedDifference:
    """文档内 Σ行金额 vs 总金额。subject_key 就是角色本身。"""
    return ExpectedDifference(
        Scope.CALCULATION,
        role.value,
        "grand_total",
        DifferenceType.CALCULATION_ERROR,
        severity,
    )


def _plans(
    overrides: Mapping[DocumentRole, DocOverride] | None = None,
    omit_roles: Sequence[DocumentRole] = (),
) -> tuple[DocumentPlan, ...]:
    """按贸易链条顺序生成参与本组的单据。

    `omit_roles` 产出「只有 PO+PI」这类少角色变体（SPEC §1.3：不强制三份齐全）。
    """
    plans: list[DocumentPlan] = []
    for role in CHAIN_ORDER:
        if role in omit_roles:
            continue
        override = (overrides or {}).get(role, DocOverride())
        base = DocumentPlan(role=role)
        plans.append(
            DocumentPlan(
                role=role,
                items=override.items if override.items is not None else base.items,
                grand_total=(
                    override.grand_total if override.grand_total is not None else base.grand_total
                ),
                currency=override.currency if override.currency is not None else base.currency,
                incoterm=override.incoterm if override.incoterm is not None else base.incoterm,
                payment_terms=(
                    override.payment_terms
                    if override.payment_terms is not None
                    else base.payment_terms
                ),
            )
        )
    return tuple(plans)


_Q = DocumentRole.QUOTATION
_P = DocumentRole.PURCHASE_ORDER
_I = DocumentRole.PROFORMA_INVOICE


# --------------------------------------------------------------------------------
# 12 组语义用例 + 1 组两文件变体（SPEC §16.2）。版面变体见 build_cases()。
#
# 基准单据（三份内容逐字相同）本身产出 **0 条差异**，推导如下：
#   文档级：Ref No. / Date / Buyer / Seller / Currency / Trade Terms（term+地点+版本）
#           / Payment Terms / Delivery Time / Grand Total 全部相等；
#           destination / shipping_method / remarks 三份都没有 -> 不是差异
#           （SPEC §9.8：全都没有的字段不产出 MISSING_VALUE）。
#   行项目：三个 SKU 各三份一一对齐（FULL + UNIQUE），description / quantity /
#           unit / unit_price / line_total 全等；其余字段三份都没有。
#   文档内：1000×1.25=1250、500×2.40=1200、200×3.10=620，Σ=3070=Grand Total。
# 因此下面每一组的 total_differences = 0 + 该组植入差异所必然引发的条数。
# --------------------------------------------------------------------------------


def _semantic_cases() -> tuple[CasePlan, ...]:
    """12 组语义用例 + 1 组两文件变体。**期望值全部手工推导，不许倒推。**"""
    return (
        # ---------------------------------------------------------------- 1
        CasePlan(
            case_id="identical",
            description="三份单据内容逐字一致：期望零差异，CRITICAL 误报必须为 0",
            documents=_plans(),
            required_differences=(),
            # 基准 0 条，未植入任何差异。
            total_differences=0,
        ),
        # ---------------------------------------------------------------- 2
        CasePlan(
            case_id="po_quantity_changed",
            description="客户 PO 把 AB-100 从 1000 改成 1200（PO 自身金额已同步）",
            documents=_plans(
                {
                    _P: DocOverride(
                        items=(
                            LineItem("AB-100", "Ceramic Mug 350ml", 1200, "PCS", "1.25", "1500.00"),
                            PLATE,
                            BOWL,
                        ),
                        # 1500.00 + 1200.00 + 620.00
                        grand_total="3320.00",
                    )
                }
            ),
            required_differences=(
                # PO 与 PI 数量不同 -> ORDER_TO_CONFIRMATION -> CRITICAL
                _line_diff("AB-100", "quantity", DifferenceType.VALUE_CONFLICT, Severity.CRITICAL),
                # 数量变了金额必然跟着变，同样落在 PO→PI 段
                _line_diff(
                    "AB-100", "line_total", DifferenceType.VALUE_CONFLICT, Severity.CRITICAL
                ),
                _doc_diff("grand_total", DifferenceType.VALUE_CONFLICT, Severity.CRITICAL),
            ),
            # 3 条 = 上面三条；PO 内部算术自洽，不产出 CALCULATION_ERROR。
            total_differences=3,
        ),
        # ---------------------------------------------------------------- 3
        CasePlan(
            case_id="pi_unit_price_wrong",
            description="PI 把 AB-200 单价改成 2.50 却没改行金额（真实世界最常见的一种错）",
            documents=_plans(
                {
                    _I: DocOverride(
                        items=(
                            MUG,
                            LineItem("AB-200", "Ceramic Plate 8in", 500, "PCS", "2.50", "1200.00"),
                            BOWL,
                        )
                    )
                }
            ),
            required_differences=(
                _line_diff(
                    "AB-200", "unit_price", DifferenceType.VALUE_CONFLICT, Severity.CRITICAL
                ),
                # 500 × 2.50 = 1250，表上写 1200 -> 文档内恒等式失败，恒为 CRITICAL
                _calc_line_diff(_I, "AB-200", Severity.CRITICAL),
            ),
            # 2 条 = 上面两条。行金额三份都是 1200.00 -> 无跨文档 line_total 冲突；
            # PI 的 Σ行金额仍是 3070.00 = Grand Total -> 无未解释差额。
            total_differences=2,
        ),
        # ---------------------------------------------------------------- 4
        CasePlan(
            case_id="currency_mismatch",
            description="PI 币种写成 EUR，与 Q/PO 的 USD 不一致：金额结构上无法比较",
            documents=_plans({_I: DocOverride(currency=CURRENCY_EUR)}),
            required_differences=(
                _doc_diff("currency", DifferenceType.VALUE_CONFLICT, Severity.CRITICAL),
                # 币种不同 -> 金额一律 INCOMPARABLE（REVIEW），
                # **绝不能坍缩成 VALUE_CONFLICT（假警报）或不产出（危险的沉默）**
                _doc_diff("grand_total", DifferenceType.INCOMPARABLE, Severity.REVIEW),
                _line_diff("AB-100", "unit_price", DifferenceType.INCOMPARABLE, Severity.REVIEW),
                _line_diff("AB-100", "line_total", DifferenceType.INCOMPARABLE, Severity.REVIEW),
                _line_diff("AB-200", "unit_price", DifferenceType.INCOMPARABLE, Severity.REVIEW),
                _line_diff("AB-200", "line_total", DifferenceType.INCOMPARABLE, Severity.REVIEW),
                _line_diff("AB-300", "unit_price", DifferenceType.INCOMPARABLE, Severity.REVIEW),
                _line_diff("AB-300", "line_total", DifferenceType.INCOMPARABLE, Severity.REVIEW),
            ),
            # 8 条 = 币种冲突 1 + 总金额无法比较 1 + 三个 SKU × (单价, 行金额) 6。
            # 数量与单位不受币种影响，仍然判等；各文档内部算术仍然自洽。
            total_differences=8,
        ),
        # ---------------------------------------------------------------- 5
        CasePlan(
            case_id="incoterm_mismatch",
            description="PI 的贸易术语写成 CIF，与 Q/PO 的 FOB 不一致",
            documents=_plans({_I: DocOverride(incoterm=INCOTERM_CIF_NINGBO)}),
            required_differences=(
                _doc_diff("incoterm", DifferenceType.VALUE_CONFLICT, Severity.CRITICAL),
            ),
            # 1 条：地点仍是 Ningbo、版本仍是 2020，只有 term 段不同。
            total_differences=1,
        ),
        # ---------------------------------------------------------------- 6
        CasePlan(
            case_id="incoterm_place_mismatch",
            description="贸易术语同为 FOB 但地点不同：Q/PO 是 Ningbo，PI 是 Shanghai",
            documents=_plans({_I: DocOverride(incoterm=INCOTERM_FOB_SHANGHAI)}),
            required_differences=(
                # 只比 term 会把这条静默吞掉，所以 named_place 是独立字段
                _doc_diff("incoterm_named_place", DifferenceType.VALUE_CONFLICT, Severity.CRITICAL),
            ),
            # 1 条：term 段（FOB）与版本（2020）都相同。
            total_differences=1,
        ),
        # ---------------------------------------------------------------- 7
        CasePlan(
            case_id="payment_terms_mismatch",
            description="付款比例不一致：Q/PO 为 30/70，PI 为 50/50",
            documents=_plans({_I: DocOverride(payment_terms=PAYMENT_50_50)}),
            required_differences=(
                _doc_diff("payment_terms", DifferenceType.VALUE_CONFLICT, Severity.CRITICAL),
            ),
            # 1 条：两侧都能结构化（比例合计 100%），因此是真差异而不是「待确认」。
            total_differences=1,
        ),
        # ---------------------------------------------------------------- 8
        CasePlan(
            case_id="po_extra_sku",
            description="PO 比 Q/PI 多了一个 AB-400（客户临时加订）",
            documents=_plans(
                {
                    _P: DocOverride(
                        items=(MUG, PLATE, BOWL, TRAY),
                        # 3070.00 + 1200.00
                        grand_total="4270.00",
                    )
                }
            ),
            required_differences=(
                # 只出现在 PO -> SKU 存在性有向表判 CRITICAL（我方两份单据都没有）
                _line_diff(
                    "AB-400", "internal_sku", DifferenceType.UNMATCHED_LINE_ITEM, Severity.CRITICAL
                ),
                _doc_diff("grand_total", DifferenceType.VALUE_CONFLICT, Severity.CRITICAL),
            ),
            # 2 条：AB-400 组是 ISOLATED，没有可比对象，不再产出字段差异；
            # 其余三个 SKU 三份全等；PO 内部 300×4.00=1200、Σ=4270 自洽。
            total_differences=2,
        ),
        # ---------------------------------------------------------------- 9
        CasePlan(
            case_id="pi_missing_sku",
            description="PI 漏掉了已下单的 AB-300（漏货）",
            documents=_plans(
                {
                    _I: DocOverride(
                        items=(MUG, PLATE),
                        # 1250.00 + 1200.00
                        grand_total="2450.00",
                    )
                }
            ),
            required_differences=(
                # Q+PO 有而 PI 无 -> 漏货 -> CRITICAL
                _line_diff(
                    "AB-300", "internal_sku", DifferenceType.UNMATCHED_LINE_ITEM, Severity.CRITICAL
                ),
                _doc_diff("grand_total", DifferenceType.VALUE_CONFLICT, Severity.CRITICAL),
            ),
            # 2 条：AB-300 组覆盖为 PARTIAL，Q 与 PO 之间照常比较且全等
            # （SPEC §9.5 第三行：缺席不得屏蔽已存在角色之间的真冲突）；
            # PI 的 Σ行金额 2450.00 = 自己的 Grand Total，无未解释差额。
            total_differences=2,
        ),
        # ---------------------------------------------------------------- 10
        CasePlan(
            case_id="row_order_shuffled",
            description="三份内容相同但行顺序不同：顺序不该产生任何差异",
            documents=_plans(
                {
                    _P: DocOverride(items=(BOWL, MUG, PLATE)),
                    _I: DocOverride(items=(PLATE, BOWL, MUG)),
                }
            ),
            required_differences=(),
            # 0 条：匹配走 SKU 精确匹配，line_key 与 group_key 都是自然键，
            # 与行号无关；总金额三份都是 3070.00。
            total_differences=0,
        ),
        # ---------------------------------------------------------------- 11
        CasePlan(
            case_id="duplicate_sku",
            description="PO 把 AB-100 拆成 600+400 两行：只能产出 AMBIGUOUS_MATCH 交人工",
            documents=_plans(
                {
                    _P: DocOverride(
                        items=(
                            LineItem("AB-100", "Ceramic Mug 350ml", 600, "PCS", "1.25", "750.00"),
                            LineItem("AB-100", "Ceramic Mug 350ml", 400, "PCS", "1.25", "500.00"),
                            PLATE,
                            BOWL,
                        ),
                        # 750.00 + 500.00 + 1200.00 + 620.00 = 3070.00，与另两份一致
                        grand_total=BASE_GRAND_TOTAL,
                    )
                }
            ),
            required_differences=(
                # matching 层不做求和、不做拆合推断（CLAUDE.md 红线）：
                # 多成员组一律 AMBIGUOUS_MATCH(REVIEW)，且该组**不做任何字段比较**
                _line_diff("AB-100", None, DifferenceType.AMBIGUOUS_MATCH, Severity.REVIEW),
            ),
            # 1 条：AB-100 组被多重性阻断（0 条 VALUE_CONFLICT），
            # AB-200 / AB-300 三份全等，PO 内部 600×1.25=750、400×1.25=500、Σ=3070 自洽。
            total_differences=1,
        ),
        # ---------------------------------------------------------------- 12
        CasePlan(
            case_id="line_total_wrong",
            description="PI 的 AB-200 行金额写成 1300.00（正确值 1200.00），总金额未同步",
            documents=_plans(
                {
                    _I: DocOverride(
                        items=(
                            MUG,
                            LineItem("AB-200", "Ceramic Plate 8in", 500, "PCS", "2.40", "1300.00"),
                            BOWL,
                        )
                    )
                }
            ),
            required_differences=(
                # 500 × 2.40 = 1200，表上写 1300 -> 文档内恒等式失败
                _calc_line_diff(_I, "AB-200", Severity.CRITICAL),
                # Σ行金额 3170.00 vs 总金额 3070.00：**不判 CRITICAL**，
                # 判「未解释差额」REVIEW（真实 PI 几乎总有运费/折扣/模具费）
                _calc_total_diff(_I, Severity.REVIEW),
                # 行金额与另两份不同 -> PO→PI 段 CRITICAL
                _line_diff(
                    "AB-200", "line_total", DifferenceType.VALUE_CONFLICT, Severity.CRITICAL
                ),
            ),
            # 3 条 = 上面三条。单价三份都是 2.40，总金额三份都是 3070.00。
            total_differences=3,
        ),
        # ---------------------------------------------------------------- 13
        CasePlan(
            case_id="two_docs_only",
            description="只上传 PO + PI 两份且内容一致：缺席的报价单不得产生任何差异",
            documents=_plans(omit_roles=(_Q,)),
            required_differences=(),
            # 0 条：未上传的角色不进入比较集合，因而不产生任何 MISSING_VALUE
            # （SPEC §9.8；否则两文件场景会被几十条假缺失淹没）。
            total_differences=0,
        ),
    )


def _layout_variant(
    *, case_id: str, description: str, source: CasePlan, layout: Layout
) -> CasePlan:
    """版面变体组（SPEC §16.1）。

    只换版面、**不换任何语义**，因此期望值直接沿用被复用组的
    `required_differences` 与 `total_differences`（SPEC §16.1「复用其他组的
    expected.json」）。这是单一真源：改了源组的期望，变体自动跟着改，
    两份期望不可能各写一遍然后漂移。

    这也是断言的全部力量所在——**换版面后差异集合必须逐条不变**。
    多一条是版面噪音制造的误报，少一条是版面变化造成的漏报。
    """
    return CasePlan(
        case_id=case_id,
        description=description,
        documents=tuple(replace(doc, layout=layout) for doc in source.documents),
        required_differences=source.required_differences,
        total_differences=source.total_differences,
    )


def build_cases() -> tuple[CasePlan, ...]:
    """全部用例：12 组语义 + 1 组两文件变体 + 3 组版面变体（SPEC §16.1、§16.2）。"""
    semantic = _semantic_cases()
    by_id = {case.case_id: case for case in semantic}
    variants = (
        _layout_variant(
            case_id="layout_merged_title",
            description="版面变体：公司抬头 + 合并大标题；语义同 identical，期望仍是零差异",
            # 复用 identical：合并单元格与抬头是表头打分器最容易被带偏的地方，
            # 一旦被带偏就会整片误报，所以拿「零差异」那一组来压最狠。
            source=by_id["identical"],
            layout=LAYOUT_MERGED_TITLE,
        ),
        _layout_variant(
            case_id="layout_chinese_headers",
            description="版面变体：纯中文标签与表头；语义同 payment_terms_mismatch",
            # 复用 payment_terms_mismatch：它的必命中差异是**文档级**字段，
            # 只有中文标签真的被提取到，这条断言才可能通过。
            # 换成零差异的组就会变成危险的绿——什么都没提取到也是零差异。
            source=by_id["payment_terms_mismatch"],
            layout=LAYOUT_CHINESE,
        ),
        _layout_variant(
            case_id="layout_second_sheet",
            description="版面变体：订单表在第二个 sheet；语义同 pi_unit_price_wrong",
            # 复用 pi_unit_price_wrong：必命中差异同时覆盖跨文档字段比较与
            # 文档内算术校验，选错 sheet 会让两条一起消失。
            source=by_id["pi_unit_price_wrong"],
            layout=LAYOUT_SECOND_SHEET,
        ),
    )
    return semantic + variants


# --------------------------------------------------------------------------------
# XLSX 写出
# --------------------------------------------------------------------------------


def document_rows(plan: DocumentPlan) -> list[list[object]]:
    """一份单据的网格。

    版面刻意做成真实样子：公司抬头 -> 大标题 -> Key-Value 区 -> 行项目表 -> 合计行。
    表头不在第 1 行，正是为了让表头打分器（SPEC §6.1）真的被考到。

    标签与列头文本全部取自 `plan.layout`，行序与列序恒定 —— 版面变体只换文本，
    **不换任何语义**。
    """
    layout = plan.layout
    rows: list[list[object]] = [
        [SELLER],
        [SELLER_ADDRESS],
        [],
        [plan.title],
        [],
        [layout.ref_no, REFERENCE_NO, "", layout.date, DOC_DATE],
        [layout.buyer, BUYER, "", layout.seller, SELLER],
        [layout.currency, plan.currency, "", layout.trade_terms, plan.incoterm],
        [layout.payment_terms, plan.payment_terms],
        [layout.delivery, plan.delivery_terms],
        [],
        list(layout.headers),
    ]
    rows.extend(
        [item.sku, item.description, item.quantity, item.unit, item.unit_price, item.line_total]
        for item in plan.items
    )
    rows.append([])
    rows.append(["", "", "", "", layout.grand_total, plan.grand_total])
    return rows


def _fix_core_properties(data: bytes) -> bytes:
    """把 docProps/core.xml 里的时间戳钉死，并**校验钉住了**。

    静默失败是这里唯一真正的风险：`re.sub` 匹配不上就原样返回，时间戳退回
    `utcnow()`，而 OOXML 时间戳只到秒，同一秒内跑完的「两次生成比 sha256」照样全绿。
    所以匹配数不为 1、或替换后仍有非固定值的 dcterms 日期，一律当场抛。
    """
    fixed, replaced = _MODIFIED_XML.subn(b"\\g<1>" + _DOC_TIMESTAMP_XML + b"\\g<2>", data)
    if replaced != 1:
        raise RuntimeError(
            f"{_CORE_PROPERTIES_ENTRY} 里没有恰好一个 dcterms:modified（实际 {replaced} 个）："
            "openpyxl 的写法已变，固定时间戳失效，产物将不再可复现。"
        )
    stray = [
        match.group(2).decode("utf-8", "replace")
        for match in _ANY_DCTERMS_DATE.finditer(fixed)
        if match.group(2) != _DOC_TIMESTAMP_XML
    ]
    if stray:
        raise RuntimeError(
            f"{_CORE_PROPERTIES_ENTRY} 里仍有非固定时间戳 {stray}：这是一处新的非确定性来源。"
        )
    return fixed


def _repack_deterministically(payload: bytes) -> bytes:
    """按固定时间戳重打包 zip。

    两处时间戳必须一起清掉，否则同一份内容两次生成的 sha256 必然不同，
    SPEC §16.1「二次生成 sha256 一致」无从验证：

      1. zip 条目的 `date_time` —— openpyxl 用 `writestr(str, ...)` 保存，
         zipfile 会给每个条目写入**当前本地时间**
      2. `docProps/core.xml` 的 `dcterms:modified` —— openpyxl 在写盘那一刻
         覆盖成 `utcnow()`

    条目顺序、内容、压缩方式全部原样保留。这是可复现构建的常规做法
    （strip-nondeterminism 同一手法），改动范围仅限上述两个时间戳。
    """
    out = BytesIO()
    with (
        zipfile.ZipFile(BytesIO(payload)) as src,
        zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst,
    ):
        for info in src.infolist():
            data = src.read(info.filename)
            if info.filename == _CORE_PROPERTIES_ENTRY:
                data = _fix_core_properties(data)
            fixed = zipfile.ZipInfo(filename=info.filename, date_time=_ZIP_TIMESTAMP)
            fixed.compress_type = zipfile.ZIP_DEFLATED
            fixed.external_attr = info.external_attr
            fixed.internal_attr = info.internal_attr
            fixed.create_system = info.create_system
            dst.writestr(fixed, data)
    return out.getvalue()


def workbook_bytes(plan: DocumentPlan) -> bytes:
    """把一份单据渲染成确定性的 .xlsx 字节。"""
    layout = plan.layout
    wb = Workbook()
    worksheet = wb.active
    if worksheet is None:  # pragma: no cover - openpyxl 新建工作簿必有活动表
        raise RuntimeError("openpyxl 未能创建活动工作表")
    if layout.lead_sheet_title is not None:
        # 订单表落到第二个 sheet：前面这张说明表必须过不了表头门槛，
        # 否则测的就不是「多 sheet 里选对了表」。
        worksheet.title = layout.lead_sheet_title
        for lead_row in LEAD_SHEET_ROWS:
            worksheet.append(list(lead_row))
        worksheet = wb.create_sheet()
    worksheet.title = plan.title
    for row in document_rows(plan):
        worksheet.append(row)
    if layout.merge_title:
        for cell_range in MERGED_TITLE_RANGES:
            worksheet.merge_cells(cell_range)
    # 文档属性里的时间戳同样要固定，否则 docProps/core.xml 每次生成都不同。
    wb.properties.created = _DOC_TIMESTAMP
    wb.properties.modified = _DOC_TIMESTAMP
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return _repack_deterministically(buffer.getvalue())


def expected_json(case: CasePlan) -> str:
    """expected.json 全文。UTF-8、LF、末尾换行——跨平台字节一致。"""
    return json.dumps(case.to_json(), ensure_ascii=False, indent=2) + "\n"


def write_case(case: CasePlan, root: Path = FIXTURES_ROOT) -> tuple[Path, ...]:
    """写出一组用例的全部文件。返回写出的路径（顺序确定）。"""
    orders_dir = root / "orders" / case.case_id
    expected_dir = root / "expected"
    orders_dir.mkdir(parents=True, exist_ok=True)
    expected_dir.mkdir(parents=True, exist_ok=True)

    written: list[Path] = []
    for plan in case.documents:
        path = orders_dir / plan.filename
        path.write_bytes(workbook_bytes(plan))
        written.append(path)

    expected_path = expected_dir / f"{case.case_id}.json"
    expected_path.write_text(expected_json(case), encoding="utf-8", newline="\n")
    written.append(expected_path)
    return tuple(written)


def build_all(root: Path = FIXTURES_ROOT) -> tuple[Path, ...]:
    """生成全部用例。同一 root 重复调用必得逐字节相同的产物。"""
    written: list[Path] = []
    for case in build_cases():
        written.extend(write_case(case, root))
    return tuple(written)


def case_files(case: CasePlan, root: Path = FIXTURES_ROOT) -> tuple[Path, ...]:
    """一组用例应当存在的全部文件（golden 测试用它判断是否需要重新生成）。"""
    orders_dir = root / "orders" / case.case_id
    return (
        *(orders_dir / plan.filename for plan in case.documents),
        root / "expected" / f"{case.case_id}.json",
    )


# --------------------------------------------------------------------------------
# 入口
# --------------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="生成 MVP-0 的 golden fixtures（12 语义 + 1 两文件变体 + 3 版面变体）"
    )
    parser.add_argument("--list", action="store_true", help="只列出用例，不写任何文件")
    parser.add_argument(
        "--root",
        type=Path,
        default=FIXTURES_ROOT,
        help=f"输出根目录（默认 {FIXTURES_ROOT}）",
    )
    args = parser.parse_args(argv)

    cases = build_cases()

    if args.list:
        for case in cases:
            roles = "+".join(role.value for role in case.compared_roles)
            print(
                f"{case.case_id:<24} {roles:<48} "
                f"必命中 {len(case.required_differences)} 条 / 共 {case.total_differences} 条"
            )
        return 0

    total_files = 0
    for case in cases:
        written = write_case(case, args.root)
        total_files += len(written)
        print(f"已生成 {case.case_id}：{len(written)} 个文件")
    print(f"共 {len(cases)} 组、{total_files} 个文件，输出根目录 {args.root}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
